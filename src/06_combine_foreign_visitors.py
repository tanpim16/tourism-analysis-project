"""
Combine foreign tourist arrival data from MOTS Excel files (2024 & 2025)
into a single long-format CSV.

Source: Ministry of Tourism and Sports (MOTS), Thailand
  - Jan_Dec_2024.xlsx  →  2023 & 2024 monthly data
  - Jan_Dec_2025.xlsx  →  2024 & 2025 monthly data
"""

import pandas as pd
import openpyxl
import os

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# Rows that are region/sub-region aggregates (not individual countries)
AGGREGATE_ROWS = {
    'Asia and the Pacific', 'South-East Asia', 'North-East Asia',
    'South Asia', 'Oceania', 'Europe', 'Northern Europe',
    'Western Europe', 'Central/Eastern Europe', 'Southern/Medit. Europe',
    'America', 'Middle East', 'Africa', 'Grand Total',
}

# ─── Region mapping (parent region for each sub-block) ────────────────────────
REGION_STARTS = [
    (6,  'Asia and the Pacific'),
    (36, 'Oceania'),
    (40, 'Europe'),
    (85, 'America'),
    (96, 'Middle East'),
    (109, 'Africa'),
]


def get_region(row_num):
    """Return the parent region based on row number."""
    region = 'Unknown'
    for start, name in REGION_STARTS:
        if row_num >= start:
            region = name
    return region


def parse_excel(filepath, primary_year, comparison_year):
    """
    Parse a MOTS foreign tourist arrivals Excel file.
    Returns a list of dicts with keys: year, month, country, region, is_aggregate, visitors.
    
    primary_year columns: B–M  (cols 2–13)
    comparison_year columns: O–Z  (cols 15–26)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []

    for row_num in range(6, 117):  # rows 6–116 contain data
        country = ws.cell(row_num, 1).value
        if country is None:
            continue
        country = country.strip()
        if not country:
            continue

        is_agg = country in AGGREGATE_ROWS
        region = get_region(row_num)

        # Primary year: cols 2–13
        for mi, month in enumerate(MONTHS):
            val = ws.cell(row_num, 2 + mi).value
            if val is not None:
                records.append({
                    'year': primary_year,
                    'month': month,
                    'month_year': f'{month}-{primary_year}',
                    'country': country,
                    'region': region,
                    'is_aggregate': is_agg,
                    'visitors': int(val) if val else 0,
                })

        # Comparison year: cols 15–26
        for mi, month in enumerate(MONTHS):
            val = ws.cell(row_num, 15 + mi).value
            if val is not None:
                records.append({
                    'year': comparison_year,
                    'month': month,
                    'month_year': f'{month}-{comparison_year}',
                    'country': country,
                    'region': region,
                    'is_aggregate': is_agg,
                    'visitors': int(val) if val else 0,
                })

    return records


# ─── Parse both files ─────────────────────────────────────────────────────────
base = 'ImportData/foreigner_data'

print('Parsing Jan_Dec_2024.xlsx (2024 + 2023) ...')
records_2024 = parse_excel(f'{base}/Jan_Dec_2024.xlsx',
                           primary_year=2024, comparison_year=2023)

print('Parsing Jan_Dec_2025.xlsx (2025 + 2024) ...')
records_2025 = parse_excel(f'{base}/Jan_Dec_2025.xlsx',
                           primary_year=2025, comparison_year=2024)

# ─── Combine & deduplicate ────────────────────────────────────────────────────
df = pd.DataFrame(records_2024 + records_2025)

# The 2025 file also contains 2024 data (as comparison). Deduplicate by
# keeping the version from the primary-year file (2024 file → 2024 data).
df = df.drop_duplicates(subset=['year', 'month', 'country'], keep='first')

# Sort
month_order = {m: i for i, m in enumerate(MONTHS)}
df['_month_sort'] = df['month'].map(month_order)
df = df.sort_values(['year', '_month_sort', 'country']).drop(columns='_month_sort')

print(f'\nCombined: {len(df)} rows, years {df["year"].unique()}, '
      f'{df["country"].nunique()} countries/regions')

# ─── Save ─────────────────────────────────────────────────────────────────────
out_dir = 'data/processed'
os.makedirs(out_dir, exist_ok=True)
out_path = f'{out_dir}/foreign_visitors_combined.csv'

df.to_csv(out_path, index=False)
print(f'✓ Saved: {out_path}')
print(df.head(10).to_string(index=False))
